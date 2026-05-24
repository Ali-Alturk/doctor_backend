"""
PDF and Excel export generators for shift schedules.
"""

import io
import calendar
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import (
    db, Shift, ShiftAssignment, Doctor, MonthlySchedule,
)
from utils.fairness import compute_fairness


# =========================================================================
# PDF EXPORT
# =========================================================================

def generate_pdf(schedule_id, db_session):
    """Generate a landscape A4 PDF with a monthly calendar grid layout."""
    try:
        schedule = db_session.query(MonthlySchedule).get(schedule_id)
        if schedule is None:
            raise ValueError("Schedule not found")

        shifts = (
            db_session.query(Shift)
            .filter_by(schedule_id=schedule_id)
            .order_by(Shift.date)
            .all()
        )

        # Build shift data keyed by date
        shift_data = {}
        for shift in shifts:
            assignments = (
                db_session.query(ShiftAssignment)
                .filter_by(shift_id=shift.id)
                .all()
            )
            doctors_info = []
            for a in assignments:
                doc = db_session.query(Doctor).get(a.doctor_id)
                if doc:
                    doctors_info.append({
                        "name": doc.full_name,
                        "seniority": doc.seniority_level,
                        "is_manual": a.is_manual_override,
                    })
            shift_data[shift.date] = {
                "day_type": shift.day_type,
                "attending_name": shift.attending_name or "",
                "attending_degree": shift.attending_degree or "",
                "doctors": doctors_info,
            }

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=15 * mm,
            rightMargin=15 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = ParagraphStyle(
            "Title",
            parent=styles["Heading1"],
            alignment=TA_CENTER,
            fontSize=16,
            spaceAfter=12,
        )
        month_name = calendar.month_name[schedule.month]
        elements.append(
            Paragraph(
                f"On-Call Schedule — {month_name} {schedule.year}",
                title_style,
            )
        )
        elements.append(Spacer(1, 10))

        cell_style = ParagraphStyle(
            "Cell",
            parent=styles["Normal"],
            fontSize=6,
            leading=8,
        )
        header_style = ParagraphStyle(
            "Header",
            parent=styles["Normal"],
            fontSize=7,
            leading=9,
            textColor=colors.white,
        )

        # Build calendar grid
        cal = calendar.Calendar(firstweekday=6)  # Sunday start
        weeks = cal.monthdatescalendar(schedule.year, schedule.month)
        day_names = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]

        header_row = [Paragraph(d, header_style) for d in day_names]
        table_data = [header_row]

        DAY_TYPE_COLORS = {
            "workday": colors.HexColor("#FFFFFF"),
            "weekend": colors.HexColor("#FFF3CD"),
            "holiday": colors.HexColor("#F8D7DA"),
        }

        cell_bg_map = {}

        for week_idx, week in enumerate(weeks):
            row = []
            for day_idx, day in enumerate(week):
                if day.month != schedule.month:
                    row.append(Paragraph("", cell_style))
                    continue

                data = shift_data.get(day)
                if data is None:
                    row.append(Paragraph(str(day.day), cell_style))
                    continue

                # Build cell content
                type_indicator = ""
                if data["day_type"] == "weekend":
                    type_indicator = " [W]"
                elif data["day_type"] == "holiday":
                    type_indicator = " [H]"

                lines = [f"<b>{day.day}{type_indicator}</b>"]
                if data["attending_name"]:
                    degree_abbr = (
                        "Prof." if data["attending_degree"] == "Professor"
                        else "Spec." if data["attending_degree"] == "Specialist"
                        else ""
                    )
                    lines.append(f"<i>Att: {degree_abbr} {data['attending_name']}</i>")

                for doc_info in data["doctors"]:
                    level_abbr = doc_info["seniority"][0]  # S, M, J
                    manual = " *" if doc_info["is_manual"] else ""
                    lines.append(f"[{level_abbr}] {doc_info['name']}{manual}")

                content = "<br/>".join(lines)
                row.append(Paragraph(content, cell_style))

                # Track bg color
                bg_color = DAY_TYPE_COLORS.get(data["day_type"], colors.white)
                cell_bg_map[(week_idx + 1, day_idx)] = bg_color

            table_data.append(row)

        col_width = (landscape(A4)[0] - 30 * mm) / 7
        table = Table(
            table_data,
            colWidths=[col_width] * 7,
            repeatRows=1,
        )

        style_commands = [
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]

        for (row, col), bg_color in cell_bg_map.items():
            style_commands.append(("BACKGROUND", (col, row), (col, row), bg_color))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)

        # Legend
        elements.append(Spacer(1, 10))
        legend_style = ParagraphStyle(
            "Legend", parent=styles["Normal"], fontSize=7,
        )
        elements.append(
            Paragraph(
                "[W] = Weekend &nbsp; [H] = Holiday &nbsp; "
                "[S] = Senior &nbsp; [M] = Mid &nbsp; [J] = Junior &nbsp; "
                "* = Manual Override",
                legend_style,
            )
        )

        doc.build(elements)
        return buffer.getvalue()

    except Exception as e:
        raise RuntimeError(f"PDF generation error: {str(e)}")


# =========================================================================
# EXCEL EXPORT
# =========================================================================

def generate_excel(schedule_id, db_session):
    """Generate a multi-sheet Excel file with schedule and fairness data."""
    try:
        schedule = db_session.query(MonthlySchedule).get(schedule_id)
        if schedule is None:
            raise ValueError("Schedule not found")

        shifts = (
            db_session.query(Shift)
            .filter_by(schedule_id=schedule_id)
            .order_by(Shift.date)
            .all()
        )

        wb = openpyxl.Workbook()

        # === Sheet 1: Monthly Schedule ===
        ws1 = wb.active
        ws1.title = "Monthly Schedule"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2C3E50", end_color="2C3E50", fill_type="solid"
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "Date", "Day", "Day Type", "Attending", "Degree",
            "Doctor 1", "Doctor 2", "Doctor 3", "Doctor 4",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        TYPE_FILLS = {
            "workday": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
            "weekend": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "holiday": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        }

        for row_idx, shift in enumerate(shifts, 2):
            assignments = (
                db_session.query(ShiftAssignment)
                .filter_by(shift_id=shift.id)
                .all()
            )
            doc_names = []
            for a in assignments:
                doc = db_session.query(Doctor).get(a.doctor_id)
                if doc:
                    label = f"{doc.full_name} ({doc.seniority_level[0]})"
                    if a.is_manual_override:
                        label += " *"
                    doc_names.append(label)

            row_data = [
                shift.date.isoformat(),
                shift.date.strftime("%A"),
                shift.day_type.capitalize(),
                shift.attending_name or "",
                shift.attending_degree or "",
            ]
            row_data.extend(doc_names)
            # Pad to 4 doctors
            while len(row_data) < 9:
                row_data.append("")

            fill = TYPE_FILLS.get(shift.day_type, TYPE_FILLS["workday"])

            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Auto-fit column widths
        for col in ws1.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 4, 25)
            ws1.column_dimensions[col[0].column_letter].width = adjusted_width

        # === Sheet 2: Fairness Stats ===
        ws2 = wb.create_sheet("Fairness Stats")
        fairness_data = compute_fairness(schedule_id, db_session)

        fairness_headers = [
            "Doctor", "Seniority", "Total Shifts", "Target",
            "Delta", "Weekdays (Mon-Thu)", "Fridays", "Weekends (Sat-Sun)", "Holiday Shifts",
            "Consecutive", "Approved Leaves",
        ]
        for col_idx, header in enumerate(fairness_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, doc_stats in enumerate(fairness_data.get("by_doctor", []), 2):
            values = [
                doc_stats["doctor_name"],
                doc_stats["seniority"],
                doc_stats["total_shifts"],
                doc_stats["target_shifts"],
                doc_stats["delta"],
                doc_stats.get("weekday_shifts", 0),
                doc_stats.get("friday_shifts", 0),
                doc_stats.get("saturday_sunday_shifts", 0),
                doc_stats["holiday_shifts"],
                doc_stats["consecutive_occurrences"],
                doc_stats["approved_leaves"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Color-code delta
                if col_idx == 5:
                    if value == 0:
                        cell.fill = PatternFill(
                            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
                        )
                    elif abs(value) == 1:
                        cell.fill = PatternFill(
                            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                        )
                    elif abs(value) >= 2:
                        cell.fill = PatternFill(
                            start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
                        )

        for col in ws2.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 4, 25)
            ws2.column_dimensions[col[0].column_letter].width = adjusted_width

        # === Sheet 3: Seniority Summary ===
        ws3 = wb.create_sheet("Seniority Summary")
        sen_headers = [
            "Seniority", "Avg Weekend+Holiday", "Max", "Min", "Imbalance",
        ]
        for col_idx, header in enumerate(sen_headers, 1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        by_seniority = fairness_data.get("by_seniority", {})
        for row_idx, (level, stats) in enumerate(by_seniority.items(), 2):
            values = [
                level,
                stats["avg_weekend_shifts"],
                stats["max_weekend_shifts"],
                stats["min_weekend_shifts"],
                "YES" if stats["imbalance_flag"] else "No",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if col_idx == 5 and value == "YES":
                    cell.fill = PatternFill(
                        start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
                    )

        for col in ws3.columns:
            max_length = 0
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = min(max_length + 4, 25)
            ws3.column_dimensions[col[0].column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    except Exception as e:
        raise RuntimeError(f"Excel generation error: {str(e)}")
